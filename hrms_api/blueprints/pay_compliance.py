from __future__ import annotations
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Dict, Optional, Tuple, List

from flask import Blueprint, request, jsonify, current_app, send_file
from sqlalchemy import or_
from sqlalchemy.sql.sqltypes import String, Date, Boolean, JSON as _JSON
from hrms_api.extensions import db
from hrms_api.common.auth import requires_perms, requires_roles

from hrms_api.models.payroll.pay_run import PayRun, PayRunItem, PayRunItemLine
from hrms_api.models.payroll.components import SalaryComponent
from hrms_api.models.employee import Employee
# Use StatConfig as the configuration table for statutory settings (PF/ESI/PT/LWF)
from hrms_api.models.payroll.stat_config import StatConfig as ConfigModel  # mapping auto-detects column names
from sqlalchemy.sql.sqltypes import JSON as _JSON
from hrms_api.services.compliance_scope import resolve_configs as resolve_v2
from hrms_api.blueprints.attendance_rollup import compute_rollup as attendance_rollup
from io import BytesIO
from openpyxl import Workbook

bp = Blueprint("pay_compliance", __name__, url_prefix="/api/v1/compliance")

# ---------- tiny helpers ----------
def _ok(data=None, status=200, **meta):
    payload = {"success": True, "data": data}
    if meta: payload["meta"] = meta
    return jsonify(payload), status

def _fail(message, status=400, code=None, extra=None):
    payload = {"success": False, "error": {"message": message}}
    if code: payload["error"]["code"] = code
    if extra is not None: payload["error"]["extra"] = extra
    return jsonify(payload), status

def _d(s) -> Optional[date]:
    if not s: return None
    try: return date.fromisoformat(str(s))
    except Exception: return None

def _dec(x) -> Optional[Decimal]:
    if x is None or x == "": return None
    try: return Decimal(str(x))
    except Exception: return None

def _as_float(x):
    try: return float(x) if x is not None else None
    except Exception: return None

def _round_nearest_int(x: Decimal) -> int:
    return int((x.quantize(Decimal('1'))))

# ---- central stat snapshot (normalized) ----
def _resolve_stat_snapshot(company_id: int, state: str, as_of: date) -> Dict[str, Any]:
    """
    Return normalized PF/ESI/PT dict resolved via v2 resolver lists.
    Prefers company+state, then state, then company, then global (handled by resolve_v2 order).
    Shape:
      {"PF": {emp_rate, er_epf_rate, er_eps_rate, base_tag, wage_cap?},
       "ESI": {emp_rate, er_rate, threshold?, entry_rule},
       "PT": {state, slabs: [...]}}
    """
    state = (state or "MH").upper()
    snap: Dict[str, Any] = {"PF": None, "ESI": None, "PT": None}
    pf_list = resolve_v2("PF", company_id, state, as_of)
    esi_list = resolve_v2("ESI", company_id, state, as_of)
    pt_list = resolve_v2("PT", company_id, state, as_of)
    pf = pf_list[0] if pf_list else None
    esi = esi_list[0] if esi_list else None
    pt = pt_list[0] if pt_list else None
    if pf:
        j = getattr(pf, 'value_json', None) or {}
        snap["PF"] = {
            "base_tag": j.get("base_tag") or "BASIC_DA",
            "emp_rate": j.get("emp_rate", 0),
            "er_epf_rate": j.get("er_epf_rate", j.get("er_epf", 0)),
            "er_eps_rate": j.get("er_eps_rate", j.get("er_eps", 0)),
            "wage_cap": j.get("wage_cap"),
        }
    if esi:
        j = getattr(esi, 'value_json', None) or {}
        snap["ESI"] = {
            "emp_rate": j.get("emp_rate", 0),
            "er_rate": j.get("er_rate", 0),
            "threshold": j.get("threshold"),
            "entry_rule": j.get("entry_rule", "period_locking"),
        }
    if pt:
        j = getattr(pt, 'value_json', None) or {}
        snap["PT"] = {
            "state": state,
            "slabs": j.get("slabs") or [],
        }
    return snap

# ---------- auto mapping to your column names ----------
def _cols():
    # list of (name, column) from the mapped table
    return list(ConfigModel.__table__.columns.items())

def _pick_by_keywords_and_type(keywords: List[str], prefer_type=None, fallback_type=None):
    cols = _cols()
    # 1) keyword + type match
    for name, col in cols:
        nm = name.lower()
        if any(k in nm for k in keywords):
            if prefer_type is None or isinstance(col.type, prefer_type):
                return name
    # 2) keyword only
    for name, col in cols:
        nm = name.lower()
        if any(k in nm for k in keywords):
            return name
    # 3) bare type guess
    if prefer_type:
        for name, col in cols:
            if isinstance(col.type, prefer_type):
                return name
    if fallback_type:
        for name, col in cols:
            if isinstance(col.type, fallback_type):
                return name
    return None

# Try to detect sensible defaults:
CODE_F   = _pick_by_keywords_and_type(["code", "key", "name", "type"], prefer_type=String) or "code"
SCOPE_F  = _pick_by_keywords_and_type(["scope", "state", "region", "company"], prefer_type=String) or "scope"
VALUE_F  = _pick_by_keywords_and_type(["value", "json", "data", "payload"], prefer_type=_JSON) or "value"
FROM_F   = _pick_by_keywords_and_type(["from", "start", "begin", "valid"], prefer_type=Date, fallback_type=Date) or "effective_from"
TO_F     = _pick_by_keywords_and_type(["to", "end", "till", "validto"], prefer_type=Date, fallback_type=Date) or "effective_to"
ACTIVE_F = _pick_by_keywords_and_type(["active", "enabled", "is_active"], prefer_type=Boolean)

def _get(obj, field: Optional[str]):
    return getattr(obj, field) if (obj is not None and field and hasattr(obj, field)) else None

def _set_kw(field: Optional[str], value: Any) -> Dict[str, Any]:
    return {field: value} if field else {}

@bp.get("/_introspect")
@requires_perms("payroll.compliance.read")
def _introspect():
    # If 'company_id' is provided, return snapshot+raw for debug
    if request.args.get("company_id"):
        try:
            cid = int(request.args.get("company_id") or 0)
        except Exception:
            return _fail("company_id must be integer", 422)
        state = (request.args.get("state") or "MH").upper()
        as_of = _d(request.args.get("as_of")) or date.today()
        snap = _resolve_stat_snapshot(cid, state, as_of)
        def row(x: ConfigModel):
            if not x: return None
            return {
                "id": x.id,
                "type": getattr(x, 'type', None),
                "scope_company_id": getattr(x, 'scope_company_id', None),
                "scope_state": getattr(x, 'scope_state', None),
                "priority": getattr(x, 'priority', None),
                "effective_from": getattr(x, 'effective_from', None).isoformat() if getattr(x, 'effective_from', None) else None,
                "effective_to": getattr(x, 'effective_to', None).isoformat() if getattr(x, 'effective_to', None) else None,
                "value": getattr(x, 'value_json', None),
            }
        raw = {
            "pf": [row(x) for x in resolve_v2("PF", cid, state, as_of)],
            "esi": [row(x) for x in resolve_v2("ESI", cid, state, as_of)],
            "pt": [row(x) for x in resolve_v2("PT", cid, state, as_of)],
        }
        return _ok({"snapshot": snap, "raw": raw})

    # If 'type' is provided, perform resolver introspection stack instead of column mapping
    tp = (request.args.get("type") or "").strip().upper()
    if tp:
        company_id = request.args.get("company_id")
        try:
            company_id = int(company_id) if company_id else None
        except Exception:
            return _fail("company_id must be integer", 422)
        state = (request.args.get("state") or "MH").upper()
        on = _d(request.args.get("on")) or date.today()
        stack = resolve_v2(tp, company_id, state, on)
        def row(x: ConfigModel):
            if not x: return None
            # classify tier for transparency
            if x.scope_company_id and x.scope_state:
                tier = "company+state"
            elif x.scope_state and not x.scope_company_id:
                tier = "state"
            elif x.scope_company_id and not x.scope_state:
                tier = "company"
            else:
                tier = "global"
            return {
                "id": x.id,
                "type": getattr(x, 'type', None),
                "scope_company_id": getattr(x, 'scope_company_id', None),
                "scope_state": getattr(x, 'scope_state', None),
                "priority": getattr(x, 'priority', None),
                "effective_from": getattr(x, 'effective_from', None).isoformat() if getattr(x, 'effective_from', None) else None,
                "effective_to": getattr(x, 'effective_to', None).isoformat() if getattr(x, 'effective_to', None) else None,
                "value": getattr(x, 'value_json', None),
                "tier": tier,
            }
        return _ok({
            "type": tp, "company_id": company_id, "state": state, "on": on.isoformat(),
            "resolution_stack": [row(x) for x in stack]
        })

    # Column mapping introspect (legacy/dev)
    cols = [{"name": n, "type": str(c.type)} for n, c in _cols()]
    return _ok({
        "table": ConfigModel.__table__.name,
        "columns": cols,
        "mapping": {
            "CODE_F": CODE_F, "SCOPE_F": SCOPE_F, "VALUE_F": VALUE_F,
            "FROM_F": FROM_F, "TO_F": TO_F, "ACTIVE_F": ACTIVE_F
        }
    })

@bp.get("/configs")
@requires_perms("payroll.compliance.read")
def list_configs_admin():
    """List active and upcoming v2 configs for admin/HR.
    Query: ?type=PF|ESI|PT|LWF[&state=MH]
    """
    tp = (request.args.get("type") or "").strip().upper()
    if tp not in ("PF","ESI","PT","LWF"):
        return _fail("type must be one of PF, ESI, PT, LWF", 422)
    st = (request.args.get("state") or None)
    today = date.today()
    q = ConfigModel.query.filter(ConfigModel.type == tp).filter((ConfigModel.effective_to.is_(None)) | (ConfigModel.effective_to >= today))
    if st:
        q = q.filter((ConfigModel.scope_state == st) | (ConfigModel.scope_state.is_(None)))
    q = q.order_by(ConfigModel.type.asc(), ConfigModel.scope_state.asc(), ConfigModel.scope_company_id.asc(), ConfigModel.priority.asc(), ConfigModel.effective_from.desc())
    rows = q.all()
    def row(x: ConfigModel):
        return {
            "id": x.id,
            "type": getattr(x, 'type', None),
            "scope_company_id": getattr(x, 'scope_company_id', None),
            "scope_state": getattr(x, 'scope_state', None),
            "priority": getattr(x, 'priority', None),
            "effective_from": getattr(x, 'effective_from', None).isoformat() if getattr(x, 'effective_from', None) else None,
            "effective_to": getattr(x, 'effective_to', None).isoformat() if getattr(x, 'effective_to', None) else None,
            "value": getattr(x, 'value_json', None),
        }
    return _ok([row(r) for r in rows])

@bp.get("/export")
@requires_perms("payroll.compliance.read")
def export_run():
    run_id = request.args.get("run_id")
    fmt = (request.args.get("format") or "xlsx").lower()
    includes = (request.args.get("include") or "pf,esi,pt").lower().split(",")
    if fmt != "xlsx":
        return _fail("Only format=xlsx supported", 422)
    try:
        rid = int(run_id)
    except Exception:
        return _fail("run_id is required and must be integer", 422)
    state = (request.args.get("state") or "MH").upper()

    prev = _build_preview(rid, state)

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "WORKING"
    headers = [
        "SR.NO","EMP CODE","NAME","UAN","ESIC","MOBILE NO","AADHAR NO",
        "DAYS","PAID DAYS","NCP DAYS",
        "GROSS WAGES (esic Deduct Amount)","ROUND WITH ESIC PAID AMOUNT","ESIC 4%",
        "PF WAGES","ROUND FOR EPF","PF EMPLOYEE 12%","PENSION EMPLOYER 8.33%","PF EMPLOYER 3.67%",
        "PT CALCULATION","COMPANY"
    ]
    ws.append(headers)

    def _num(x):
        try:
            return float(x) if x is not None else 0.0
        except Exception:
            return 0.0

    for row in prev.get("employees", []) or []:
        esi = row.get("esi") or {}
        pf = row.get("pf") or {}
        pt = row.get("pt") or {}
        ws.append([
            row.get("sr_no"), row.get("emp_code"), row.get("name"), row.get("uan"), row.get("esi_no"), row.get("mobile"), row.get("aadhaar"),
            _num(row.get("days_in_month")), _num(row.get("paid_days")), _num(row.get("ncp_days")),
            _num((esi.get("sheet_alias") or {}).get("GROSS_WAGES")), _num(esi.get("rounded_paid_amount")), _num((esi.get("sheet_alias") or {}).get("ESIC_4PCT_like")),
            _num(pf.get("wage_base")), _num(pf.get("rounded_for_epf")), _num(pf.get("emp_12pct")), _num(pf.get("er_eps_8_33pct")), _num(pf.get("er_epf_3_67pct")),
            _num(pt.get("amount")), row.get("company_posting"),
        ])

    # EPF helper
    if "pf" in includes:
        ws_pf = wb.create_sheet("EPF")
        ws_pf.append(["SR.NO","EMP CODE","NAME","PF WAGES","ROUND FOR EPF","PF EMPLOYEE 12%","PENSION EMPLOYER 8.33%","PF EMPLOYER 3.67%"])
        for row in prev.get("employees", []) or []:
            pf = row.get("pf") or {}
            ws_pf.append([
                row.get("sr_no"), row.get("emp_code"), row.get("name"),
                _num(pf.get("wage_base")), _num(pf.get("rounded_for_epf")), _num(pf.get("emp_12pct")), _num(pf.get("er_eps_8_33pct")), _num(pf.get("er_epf_3_67pct")),
            ])

    # ESIC helper
    if "esi" in includes:
        ws_esi = wb.create_sheet("ESIC")
        ws_esi.append(["SR.NO","EMP CODE","NAME","GROSS WAGES","ROUND WITH ESIC PAID AMOUNT","ESI EMP","ESI ER"])
        for row in prev.get("employees", []) or []:
            esi = row.get("esi") or {}
            ws_esi.append([
                row.get("sr_no"), row.get("emp_code"), row.get("name"),
                _num((esi.get("sheet_alias") or {}).get("GROSS_WAGES")), _num(esi.get("rounded_paid_amount")), _num(esi.get("emp_amount")), _num(esi.get("er_amount")),
            ])

    # PT helper
    if "pt" in includes:
        ws_pt = wb.create_sheet("PT")
        ws_pt.append(["SR.NO","EMP CODE","NAME","PT CALCULATION"])
        tot = 0.0
        for row in prev.get("employees", []) or []:
            pt = row.get("pt") or {}
            amt = _num(pt.get("amount"))
            tot += amt
            ws_pt.append([row.get("sr_no"), row.get("emp_code"), row.get("name"), amt])
        ws_pt.append([None, None, "TOTAL", tot])

    # Stream
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    r = PayRun.query.get(rid)
    yyyymm = (r.period_end or r.period_start).strftime('%Y%m') if r else date.today().strftime('%Y%m')
    filename = f"compliance_run_{rid}_{yyyymm}.xlsx"
    return send_file(bio, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)

# ---------- config readers ----------
def _cfg_query(code: str, on: date, scope: Optional[str] = None):
    q = ConfigModel.query.filter(getattr(ConfigModel, CODE_F) == code)
    if scope is not None:
        q = q.filter(getattr(ConfigModel, SCOPE_F) == scope)
    q = (q.filter(getattr(ConfigModel, FROM_F) <= on)
           .filter(or_(getattr(ConfigModel, TO_F).is_(None), getattr(ConfigModel, TO_F) >= on))
           .order_by(getattr(ConfigModel, FROM_F).desc()))
    return q

def _cfg_get(code: str, on: date, scope: Optional[str] = None) -> Optional[Dict[str, Any]]:
    rec = _cfg_query(code, on, scope).first()
    return _get(rec, VALUE_F) if rec else None

# ---------- Config endpoints ----------
@bp.get("/stat-config")
@requires_perms("payroll.compliance.read")
def stat_config_list():
    q = ConfigModel.query
    if request.args.get("code"):
        q = q.filter(getattr(ConfigModel, CODE_F) == request.args["code"])
    if request.args.get("scope"):
        q = q.filter(getattr(ConfigModel, SCOPE_F) == request.args["scope"])
    active_on = _d(request.args.get("active_on"))
    if active_on:
        q = q.filter(
            getattr(ConfigModel, FROM_F) <= active_on,
            or_(getattr(ConfigModel, TO_F).is_(None), getattr(ConfigModel, TO_F) >= active_on),
        )
    q = q.order_by(getattr(ConfigModel, CODE_F).asc(),
                   getattr(ConfigModel, SCOPE_F).asc(),
                   getattr(ConfigModel, FROM_F).desc())
    rows = q.all()

    def row(x: ConfigModel):
        return {
            "id": x.id,
            "code": _get(x, CODE_F),
            "scope": _get(x, SCOPE_F),
            "value": _get(x, VALUE_F),
            "effective_from": _get(x, FROM_F).isoformat() if _get(x, FROM_F) else None,
            "effective_to": _get(x, TO_F).isoformat() if _get(x, TO_F) else None,
            "is_active": _get(x, ACTIVE_F) if ACTIVE_F else True,
        }
    return _ok([row(r) for r in rows])

@bp.post("/stat-config")
@requires_perms("payroll.compliance.write")
def stat_config_create():
    j = request.get_json(silent=True) or {}

    # V2 flow when 'type' provided
    if (j.get("type") or "").strip():
        tp = (j.get("type") or "").strip().upper()
        if tp not in ("PF","ESI","PT","LWF"):
            return _fail("type must be one of PF, ESI, PT, LWF", 422)
        eff_from = _d(j.get("effective_from"))
        if not eff_from:
            return _fail("effective_from is required (YYYY-MM-DD)", 422)
        eff_to = _d(j.get("effective_to"))
        if eff_to and eff_to < eff_from:
            return _fail("effective_to must be >= effective_from", 422)
        comp = j.get("scope_company_id", j.get("company_id"))
        st = j.get("scope_state", j.get("state"))
        try:
            comp = int(comp) if comp is not None and str(comp).strip() != "" else None
        except Exception:
            return _fail("company_id/scope_company_id must be integer", 422)
        st = (str(st).upper() if st else None)
        prio = j.get("priority")
        try:
            prio = int(prio) if prio is not None and str(prio).strip() != "" else 100
        except Exception:
            return _fail("priority must be integer", 422)
        value = j.get("value")
        if value is None:
            return _fail("value (JSON) is required", 422)
        # Overlap for v2: type + scope_company_id/state + window
        q = (ConfigModel.query
             .filter(ConfigModel.type == tp)
             .filter(ConfigModel.effective_from <= (eff_to or date.max))
             .filter(or_(ConfigModel.effective_to.is_(None), ConfigModel.effective_to >= eff_from)))
        if comp is not None:
            q = q.filter(ConfigModel.scope_company_id == comp)
        if st is not None:
            q = q.filter(ConfigModel.scope_state == st)
        if q.first():
            return _fail("Overlapping config period for the same type/scope", 409)

        rec = ConfigModel(
            type=tp,
            scope_company_id=comp,
            scope_state=st,
            priority=prio,
            effective_from=eff_from,
            effective_to=eff_to,
            value_json=value,
            key=(j.get("code") or j.get("key") or f"STATCFG_V2_{tp}"),
            scope=("company" if comp else ("IN" if st else None)),
            company_id=comp,
            state=st,
        )
        db.session.add(rec)
        db.session.commit()
        return _ok({
            "id": rec.id,
            "type": tp,
            "scope_company_id": comp,
            "scope_state": st,
            "priority": prio,
            "effective_from": rec.effective_from.isoformat() if rec.effective_from else None,
            "effective_to": rec.effective_to.isoformat() if rec.effective_to else None,
            "value": rec.value_json,
        }, 201)

    # Legacy path (code/scope/value)
    code = (j.get("code") or "").strip().upper()
    if not code: return _fail("code is required", 422)
    eff_from = _d(j.get("effective_from"))
    if not eff_from: return _fail("effective_from is required (YYYY-MM-DD)", 422)
    eff_to = _d(j.get("effective_to"))
    if eff_to and eff_to < eff_from: return _fail("effective_to must be >= effective_from", 422)
    scope = (j.get("scope") or "").strip() or None
    value = j.get("value")
    if value is None: return _fail("value (JSON) is required", 422)

    q_ov = (
        ConfigModel.query
            .filter(getattr(ConfigModel, CODE_F) == code,
                    getattr(ConfigModel, SCOPE_F) == scope)
            .filter(getattr(ConfigModel, FROM_F) <= (eff_to or date.max))
            .filter(or_(getattr(ConfigModel, TO_F).is_(None), getattr(ConfigModel, TO_F) >= eff_from))
    )
    comp_id = j.get("company_id") or j.get("scope_company_id")
    st = j.get("state") or j.get("scope_state")
    if scope == "company" and comp_id:
        try:
            q_ov = q_ov.filter(getattr(ConfigModel, "company_id") == int(comp_id))
        except Exception:
            pass
    if scope in ("IN", "state") and st:
        q_ov = q_ov.filter(getattr(ConfigModel, "state") == str(st))
    overlap = q_ov.first()
    if overlap:
        return _fail("Overlapping config period for the same code/scope", 409)

    fields = {}
    fields.update(_set_kw(CODE_F, code))
    fields.update(_set_kw(SCOPE_F, scope))
    fields.update(_set_kw(VALUE_F, value))
    fields.update(_set_kw(FROM_F, eff_from))
    fields.update(_set_kw(TO_F, eff_to))
    if ACTIVE_F: fields.update(_set_kw(ACTIVE_F, bool(j.get("is_active", True))))

    rec = ConfigModel(**fields)
    db.session.add(rec)
    db.session.commit()

    return _ok({
        "id": rec.id,
        "code": _get(rec, CODE_F),
        "scope": _get(rec, SCOPE_F),
        "value": _get(rec, VALUE_F),
        "effective_from": _get(rec, FROM_F).isoformat() if _get(rec, FROM_F) else None,
        "effective_to": _get(rec, TO_F).isoformat() if _get(rec, TO_F) else None,
        "is_active": _get(rec, ACTIVE_F) if ACTIVE_F else True,
    }, 201)

@bp.put("/stat-config/<int:config_id>/close")
@requires_perms("payroll.compliance.write")
def stat_config_close(config_id: int):
    cur = ConfigModel.query.get_or_404(config_id)
    j = request.get_json(silent=True) or {}
    new_from = _d(j.get("new_from"))
    if not new_from: return _fail("new_from is required (YYYY-MM-DD)", 422)
    cur_from = _get(cur, FROM_F)
    if cur_from and new_from <= cur_from:
        return _fail("new_from must be after current effective_from", 422)

    setattr(cur, TO_F, new_from - timedelta(days=1))

    if "new_value" in j:
        new_scope = j.get("new_scope", _get(cur, SCOPE_F))
        # guard overlap
        ov = _cfg_query(_get(cur, CODE_F), new_from, new_scope).filter(ConfigModel.id != cur.id).first()
        if ov: return _fail("Overlapping config period for the same code/scope", 409)
        fields = {}
        fields.update(_set_kw(CODE_F, _get(cur, CODE_F)))
        fields.update(_set_kw(SCOPE_F, new_scope))
        fields.update(_set_kw(VALUE_F, j.get("new_value")))
        fields.update(_set_kw(FROM_F, new_from))
        fields.update(_set_kw(TO_F, None))
        if ACTIVE_F: fields.update(_set_kw(ACTIVE_F, bool(j.get("new_is_active", True))))
        db.session.add(ConfigModel(**fields))

    db.session.add(cur)
    db.session.commit()
    return _ok({"closed_id": cur.id})

# ---------- PF/ESI/PT/LWF compute ----------
def _pick_basic_from_components(components: List[Dict[str, Any]], wage_tag: str = "BASIC") -> Decimal:
    for c in components or []:
        if (c.get("code") or "").upper() == wage_tag.upper():
            return _dec(c.get("amount")) or Decimal("0")
    s = Decimal("0")
    for c in components or []:
        s += _dec(c.get("amount")) or Decimal("0")
    return s

def _calc_pf(item: PayRunItem, on: date, scope="IN") -> Tuple[Decimal, Decimal]:
    cfg_emp = _cfg_get("PF_RATE_EMP", on, scope)
    cfg_er  = _cfg_get("PF_RATE_ER",  on, scope)
    if not cfg_emp or not cfg_er:
        raise KeyError("PF config missing: PF_RATE_EMP and PF_RATE_ER")
    rate_emp = Decimal(str(cfg_emp.get("rate", 0)))
    rate_er  = Decimal(str(cfg_er.get("rate", 0)))
    cap = Decimal(str(cfg_emp.get("wage_cap", 0))) if cfg_emp.get("wage_cap") else None
    tag = (cfg_emp.get("wage_tag") or "BASIC").upper()
    comps = getattr(item, ITEM_COMPONENTS_F, None) if ITEM_COMPONENTS_F else None
    basic = _pick_basic_from_components(comps, tag)
    pf_wage = min(basic, cap) if cap else basic
    return (pf_wage * rate_emp).quantize(Decimal("0.01")), (pf_wage * rate_er).quantize(Decimal("0.01"))

def _calc_esi(item: PayRunItem, on: date, scope="IN") -> Tuple[Decimal, Decimal]:
    cfg = _cfg_get("ESI_RATE", on, scope)
    if not cfg: raise KeyError("ESI config missing: ESI_RATE")
    emp_r = Decimal(str(cfg.get("emp", 0)))
    er_r  = Decimal(str(cfg.get("er", 0)))
    thr   = Decimal(str(cfg.get("threshold", 0))) if cfg.get("threshold") else None
    gross = Decimal(str(getattr(item, "gross", 0) or 0))
    if thr and gross > thr: return Decimal("0.00"), Decimal("0.00")
    return (gross * emp_r).quantize(Decimal("0.01")), (gross * er_r).quantize(Decimal("0.01"))

def _calc_pt_mh(item: PayRunItem, on: date) -> Decimal:
    slabs = _cfg_get("MH_PT_SLABS", on, "MH")
    if not slabs: raise KeyError("PT config missing: MH_PT_SLABS")
    gross = Decimal(str(getattr(item, "gross", 0) or 0))
    for slab in slabs:
        mn = Decimal(str(slab.get("min", 0)))
        mx = Decimal(str(slab.get("max", 10**12)))
        amt = Decimal(str(slab.get("amount", 0)))
        if mn <= gross <= mx:
            return amt
    return Decimal("0.00")

def _calc_lwf_mh(item: PayRunItem, on: date) -> Tuple[Decimal, Decimal]:
    cfg = _cfg_get("LWF_MH", on, "MH")
    if not cfg: raise KeyError("LWF config missing: LWF_MH")
    for row in cfg:
        m = int(row.get("month", 0))
        if m == on.month or not m:
            return Decimal(str(row.get("emp", 0))), Decimal(str(row.get("er", 0)))
    return Decimal("0.00"), Decimal("0.00")

def _compose_components(base: List[Dict[str, Any]], adds: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return list(base or []) + list(adds or [])

# -------------- V2 resolution + calculators --------------
def _first_or_none(lst):
    return lst[0] if lst else None

def _resolve_cfgs_v2(company_id: Optional[int], state: str, on: date):
    snap = {
        "PF": resolve_v2("PF", company_id, state, on),
        "ESI": resolve_v2("ESI", company_id, state, on),
        "PT": resolve_v2("PT", company_id, state, on),
        # LWF planned later
    }
    # keep a simple snapshot of the exact records used (first in each list)
    used = {k: _first_or_none(v) for k, v in snap.items()}
    return snap, used

def _pick_component_wage(item: PayRunItem, base_tag: str) -> Decimal:
    comps = getattr(item, ITEM_COMPONENTS_F, None) if ITEM_COMPONENTS_F else None
    return _pick_basic_from_components(comps, base_tag or "BASIC")

def _compute_pf_v2(item: PayRunItem, pf_cfg: Any) -> Dict[str, Any]:
    if not pf_cfg: return {}
    j = pf_cfg.value_json if hasattr(pf_cfg, 'value_json') else (pf_cfg or {})
    emp_rate = Decimal(str(j.get("emp_rate", 0)))
    er_eps_rate = Decimal(str(j.get("er_eps_rate", 0)))
    er_epf_rate = Decimal(str(j.get("er_epf_rate", 0)))
    cap = Decimal(str(j.get("wage_cap", 0))) if j.get("wage_cap") else None
    base_tag = (j.get("base_tag") or "BASIC").upper()

    base = _pick_component_wage(item, base_tag)
    wage_base = min(base, cap) if cap else base
    rounded_for_epf = Decimal(str(_round_nearest_int(wage_base)))

    emp_12pct = (wage_base * emp_rate).quantize(Decimal("0.01"))
    er_eps_833 = (wage_base * er_eps_rate).quantize(Decimal("0.01"))
    er_epf_367 = (wage_base * er_epf_rate).quantize(Decimal("0.01"))

    return {
        "wage_base": _as_float(wage_base),
        "rounded_for_epf": _as_float(rounded_for_epf),
        "emp_12pct": _as_float(emp_12pct),
        "er_eps_8_33pct": _as_float(er_eps_833),
        "er_epf_3_67pct": _as_float(er_epf_367),
        "_meta": {"base_tag": base_tag},
    }

def _compute_esi_v2(item: PayRunItem, esi_cfg: Any) -> Dict[str, Any]:
    if not esi_cfg: return {}
    j = esi_cfg.value_json if hasattr(esi_cfg, 'value_json') else (esi_cfg or {})
    emp_rate = Decimal(str(j.get("emp_rate", 0)))
    er_rate = Decimal(str(j.get("er_rate", 0)))
    threshold = Decimal(str(j.get("threshold", 0))) if j.get("threshold") else None

    gross = Decimal(str(getattr(item, "gross", 0) or 0))
    if threshold and gross > threshold:
        wage_base = Decimal("0")
    else:
        wage_base = gross
    rounded_paid_amount = Decimal(str(_round_nearest_int(wage_base)))
    emp_amount = (wage_base * emp_rate).quantize(Decimal("0.01"))
    er_amount = (wage_base * er_rate).quantize(Decimal("0.01"))

    return {
        "wage_base": _as_float(wage_base),
        "rounded_paid_amount": _as_float(rounded_paid_amount),
        "emp_rate": _as_float(emp_rate),
        "er_rate": _as_float(er_rate),
        "emp_amount": _as_float(emp_amount),
        "er_amount": _as_float(er_amount),
        "sheet_alias": {
            "GROSS_WAGES": _as_float(gross),
            "ROUND_WITH_ESIC_PAID_AMOUNT": _as_float(rounded_paid_amount),
            "ESIC_4PCT_like": _as_float((gross * Decimal("0.04")).quantize(Decimal("0.01")))
        }
    }

def _compute_pt_v2(item: PayRunItem, pt_cfg: Any, state: str) -> Dict[str, Any]:
    if not pt_cfg: return {}
    j = pt_cfg.value_json if hasattr(pt_cfg, 'value_json') else (pt_cfg or {})
    if (j.get("state") or state).upper() != state.upper():
        return {}
    gross = Decimal(str(getattr(item, "gross", 0) or 0))
    amt = Decimal("0")
    for slab in j.get("slabs", []) or []:
        mn = Decimal(str(slab.get("min", 0)))
        mx = Decimal(str(slab.get("max", 10**12)))
        if mn <= gross <= mx:
            amt = Decimal(str(slab.get("amount", 0)))
            break
    return {"amount": _as_float(amt)}

def _build_preview(run_id: int, state: str) -> Dict[str, Any]:
    r = PayRun.query.get_or_404(int(run_id))
    snap = r.period_end or r.period_start

    items = PayRunItem.query.filter_by(pay_run_id=r.id).all()
    if not items:
        raise ValueError("Run has no items. Calculate the run first.")

    # Attendance rollup for days
    roll = attendance_rollup(r.company_id, r.period_start, r.period_end, [it.employee_id for it in items])

    # Resolve v2 configs (raw lists + normalized snapshot)
    cfg_lists, cfg_used = _resolve_cfgs_v2(r.company_id, state, snap)
    stat_snap = _resolve_stat_snapshot(r.company_id, state, snap)

    employees_blk: List[Dict[str, Any]] = []
    totals_new = {
        "emp_deductions": {"pf": Decimal("0"), "esi": Decimal("0"), "pt": Decimal("0"), "lwf": Decimal("0"), "all": Decimal("0")},
        "employer_costs": {"esi": Decimal("0"), "pf_eps": Decimal("0"), "pf_epf": Decimal("0"), "all": Decimal("0")},
    }

    # Legacy preview (for compatibility)
    legacy_items: List[Dict[str, Any]] = []
    legacy_totals = {"pf_emp": Decimal("0"), "pf_er": Decimal("0"), "esi_emp": Decimal("0"), "esi_er": Decimal("0"), "pt": Decimal("0"), "lwf_emp": Decimal("0"), "lwf_er": Decimal("0")}

    missing: List[Dict[str, Any]] = []

    sr = 0
    for it in items:
        sr += 1
        emp: Employee = it.employee  # joined
        att = roll.get(it.employee_id, {"days_worked": 0, "lop_days": 0, "holidays": 0, "weekly_off": 0})
        paid_days = Decimal(str(att.get("days_worked", 0) or 0))
        ncp_days = Decimal(str(att.get("lop_days", 0) or 0))
        days_in_month = _as_float(paid_days + ncp_days)

        # V2 computations using normalized snapshot
        pf_blk = _compute_pf_v2(it, stat_snap.get("PF"))
        esi_blk = _compute_esi_v2(it, stat_snap.get("ESI"))
        pt_blk = _compute_pt_v2(it, stat_snap.get("PT"), state)

        # track missing
        if not stat_snap.get("PF"):
            missing.append({"type": "PF", "state": state, "code": "PF", "effective_from": snap.isoformat()})
        if not stat_snap.get("ESI"):
            missing.append({"type": "ESI", "state": state, "code": "ESI", "effective_from": snap.isoformat()})
        if state == "MH" and not stat_snap.get("PT"):
            missing.append({"type": "PT", "state": "MH", "code": "PT", "effective_from": snap.isoformat()})

        emp_code = getattr(emp, 'code', None)
        name = (getattr(emp, 'first_name', '') or '') + ((' ' + getattr(emp, 'last_name')) if getattr(emp, 'last_name', None) else '')
        mobile = getattr(emp, 'phone', None)
        company_posting = getattr(getattr(emp, 'location', None), 'name', None) or getattr(getattr(emp, 'company', None), 'name', None)

        # Summaries
        pf_emp_amt = Decimal(str(pf_blk.get("emp_12pct", 0) or 0)) if pf_blk else Decimal("0")
        pf_eps_amt = Decimal(str(pf_blk.get("er_eps_8_33pct", 0) or 0)) if pf_blk else Decimal("0")
        pf_epf_amt = Decimal(str(pf_blk.get("er_epf_3_67pct", 0) or 0)) if pf_blk else Decimal("0")
        esi_emp_amt = Decimal(str(esi_blk.get("emp_amount", 0) or 0)) if esi_blk else Decimal("0")
        esi_er_amt = Decimal(str(esi_blk.get("er_amount", 0) or 0)) if esi_blk else Decimal("0")
        pt_amt = Decimal(str(pt_blk.get("amount", 0) or 0)) if pt_blk else Decimal("0")

        employees_blk.append({
            "sr_no": sr,
            "emp_code": emp_code,
            "name": name.strip() or None,
            "uan": None, "esi_no": None, "mobile": mobile, "aadhaar": None,
            "days_in_month": _as_float(days_in_month), "paid_days": _as_float(paid_days), "ncp_days": _as_float(ncp_days),
            "company_posting": company_posting,
            "esi": esi_blk or None,
            "pf": pf_blk or None,
            "pt": pt_blk or ( {"amount": 0.0} if state == "MH" else None ),
            "lwf": None,
            "stat_summaries": {
                "emp_deductions": {"pf": _as_float(pf_emp_amt), "esi": _as_float(esi_emp_amt), "pt": _as_float(pt_amt), "lwf": 0.0},
                "employer_costs": {"esi": _as_float(esi_er_amt), "pf_eps": _as_float(pf_eps_amt), "pf_epf": _as_float(pf_epf_amt)},
            },
        })

        # new totals
        totals_new["emp_deductions"]["pf"] += pf_emp_amt
        totals_new["emp_deductions"]["esi"] += esi_emp_amt
        totals_new["emp_deductions"]["pt"] += pt_amt
        totals_new["employer_costs"]["esi"] += esi_er_amt
        totals_new["employer_costs"]["pf_eps"] += pf_eps_amt
        totals_new["employer_costs"]["pf_epf"] += pf_epf_amt

        # legacy block for compatibility
        comp_add: List[Dict[str, Any]] = []
        if pf_emp_amt or (pf_eps_amt + pf_epf_amt):
            comp_add.append({"code": "PF_EMP", "amount": _as_float(pf_emp_amt)})
            comp_add.append({"code": "PF_ER", "amount": _as_float(pf_eps_amt + pf_epf_amt)})
            legacy_totals["pf_emp"] += pf_emp_amt; legacy_totals["pf_er"] += (pf_eps_amt + pf_epf_amt)
        if esi_emp_amt or esi_er_amt:
            comp_add.append({"code": "ESI_EMP", "amount": _as_float(esi_emp_amt)})
            comp_add.append({"code": "ESI_ER", "amount": _as_float(esi_er_amt)})
            legacy_totals["esi_emp"] += esi_emp_amt; legacy_totals["esi_er"] += esi_er_amt
        if state == "MH" and pt_amt:
            comp_add.append({"code": "PT_MH", "amount": _as_float(pt_amt)})
            legacy_totals["pt"] += pt_amt

        gross = Decimal(str(getattr(it, "gross", 0) or 0))
        emp_deds_legacy = Decimal("0")
        for c in comp_add:
            if c["code"] in ("PF_EMP","ESI_EMP","PT_MH","LWF_EMP"):
                emp_deds_legacy += Decimal(str(c["amount"]))
        preview_net = (gross - emp_deds_legacy).quantize(Decimal("0.01"))
        legacy_items.append({
            "item_id": it.id,
            "employee_id": getattr(it, "employee_id", None),
            "add_components": comp_add,
            "gross": _as_float(gross),
            "preview_net": _as_float(preview_net),
        })

    # finalize totals
    totals_new["emp_deductions"]["all"] = totals_new["emp_deductions"]["pf"] + totals_new["emp_deductions"]["esi"] + totals_new["emp_deductions"]["pt"] + totals_new["emp_deductions"]["lwf"]
    totals_new["employer_costs"]["all"] = totals_new["employer_costs"]["esi"] + totals_new["employer_costs"]["pf_eps"] + totals_new["employer_costs"]["pf_epf"]

    # config snapshot (only the first used)
    def _snap_row(x):
        if not x: return None
        return {
            "id": x.id,
            "type": getattr(x, 'type', None),
            "scope_company_id": getattr(x, 'scope_company_id', None),
            "scope_state": getattr(x, 'scope_state', None),
            "priority": getattr(x, 'priority', None),
            "effective_from": getattr(x, 'effective_from', None).isoformat() if getattr(x, 'effective_from', None) else None,
            "effective_to": getattr(x, 'effective_to', None).isoformat() if getattr(x, 'effective_to', None) else None,
            "value": getattr(x, 'value_json', None),
        }
    # back-compat raw snapshot rows in case clients depended on row shape; but primary is normalized stat_snap
    config_snapshot_raw = {k: _snap_row(v) for k, v in cfg_used.items()}

    out = {
        "employees": employees_blk,
        "totals": {"emp_deductions": {k: _as_float(v) for k, v in totals_new["emp_deductions"].items()},
                    "employer_costs": {k: _as_float(v) for k, v in totals_new["employer_costs"].items()}},
        "config_snapshot": stat_snap,
        "config_snapshot_raw": config_snapshot_raw,
        "missing_config": missing or [],
        "can_apply": len(missing) == 0,
        # legacy mirrors
        "legacy_items": legacy_items,
        "legacy_totals": {k: _as_float(v) for k, v in legacy_totals.items()},
    }
    return out

@bp.get("/preview")
@requires_perms("payroll.compliance.read")
def preview_run():
    run_id = request.args.get("run_id")
    if not run_id: return _fail("run_id is required", 422)
    state = (request.args.get("state") or "MH").upper()
    j_company = request.args.get("company_id")
    try:
        _ = int(run_id)
        _ = int(j_company) if j_company else None
    except Exception:
        return _fail("run_id/company_id must be integers", 422)
    try:
        out = _build_preview(int(run_id), state)
    except ValueError as e:
        return _fail(str(e), 422)
    # Do not block on missing; include can_apply flag
    return _ok(out)

@bp.post("/apply")
@requires_perms("payroll.compliance.write")
def apply_to_run():
    j = request.get_json(silent=True) or {}
    run_id = j.get("run_id")
    if not run_id:
        return _fail("run_id is required", 422)

    # --- helpers (local to keep this function drop-in) -----------------------
    def _coerce_types(val):
        """
        Accept: None | ["PF","ESI"] | {"PF":true,...} | "PF, ESI, PT"
        Return: set({"PF","ESI","PT","LWF"}) subset
        """
        ALL = {"PF", "ESI", "PT", "LWF"}
        if val is None:
            return set(ALL)
        if isinstance(val, dict):
            return {k.strip().upper() for k, v in val.items() if v} & ALL
        if isinstance(val, (list, tuple)):
            return {str(x).strip().upper() for x in val} & ALL
        if isinstance(val, str):
            return {t.strip().upper() for t in val.split(",")} & ALL
        return set()

    def _first_or_none(x):
        """If resolver or preview gave a list, use the first row; else return as-is."""
        if isinstance(x, (list, tuple)):
            return x[0] if x else None
        return x

    def _as_decimal(x):
        from decimal import Decimal
        return Decimal(str(x or 0))

    # Normalize optional filters (not strictly required by older code,
    # but prevents crashes if upstream starts honoring them)
    wanted_types = _coerce_types(j.get("types"))
    state = (j.get("state") or "MH").upper()

    r = PayRun.query.get_or_404(int(run_id))
    if r.status == "locked":
        return _fail("Run is locked; cannot apply compliance.", 409)
    if r.status not in ("calculated", "approved"):
        return _fail(
            f"Run in status '{r.status}' cannot apply compliance. Recalculate/approve first.",
            409,
        )

    # Reuse preview logic
    try:
        prev = _build_preview(int(run_id), state)
    except ValueError as e:
        return _fail(str(e), 422)

    # If preview says we can't apply, explain which configs are missing
    if not (prev.get("can_apply", True)):
        return _fail("Missing statutory config; cannot apply.", 422, extra=prev)

    # --- Ensure components exist (idempotent) --------------------------------
    def _ensure_component(code: str, name: str, ctype: str) -> SalaryComponent:
        sc = SalaryComponent.query.filter_by(code=code).first()
        if sc:
            return sc
        sc = SalaryComponent(code=code, name=name, type=ctype)
        db.session.add(sc)
        db.session.flush()
        return sc

    comp_pf_emp = _ensure_component("PF_EMP", "PF Employee", "deduction")
    comp_pf_eps = _ensure_component("PF_ER_EPS", "PF Employer EPS", "earning")
    comp_pf_epf = _ensure_component("PF_ER_EPF", "PF Employer EPF", "earning")
    comp_esi_emp = _ensure_component("ESI_EMP", "ESI Employee", "deduction")
    comp_esi_er = _ensure_component("ESI_ER", "ESI Employer", "earning")
    comp_pt = _ensure_component("PT", "Professional Tax", "deduction")

    applied = 0

    # Legacy preview data
    legacy_items = prev.get("legacy_items", []) or []
    legacy_map = {li.get("item_id"): li for li in legacy_items}

    # config_snapshot can be dict ({"PF": {...}, "ESI": {...}}) OR a list of rows
    cfg_ids = {}
    cfg_snapshot = prev.get("config_snapshot")
    if isinstance(cfg_snapshot, dict):
        for k, v in cfg_snapshot.items():
            cfg_ids[k] = (v.get("id") if isinstance(v, dict) else getattr(v, "id", None))
    elif isinstance(cfg_snapshot, (list, tuple)):
        for row in cfg_snapshot:
            # accept variations: row may be dict or ORM; expect 'type'/'code'
            t = (row.get("type") if isinstance(row, dict) else getattr(row, "type", None)) or \
                (row.get("code") if isinstance(row, dict) else getattr(row, "code", None))
            if t:
                cfg_ids[str(t).upper()] = (
                    row.get("id") if isinstance(row, dict) else getattr(row, "id", None)
                )

    for it in PayRunItem.query.filter_by(pay_run_id=r.id).all():
        li = legacy_map.get(it.id)
        if li is None:
            continue

        # Gather computed statutory amounts from legacy/add_components
        pf_emp_amt = _as_decimal(0)
        pf_eps_amt = _as_decimal(0)
        pf_epf_amt = _as_decimal(0)
        esi_emp_amt = _as_decimal(0)
        esi_er_amt = _as_decimal(0)
        pt_amt = _as_decimal(0)

        add_components = (li.get("add_components") or []) if isinstance(li, dict) else []
        for c in add_components:
            code = (c.get("code") or "").upper()
            amt = _as_decimal(c.get("amount", 0))

            # Respect requested types (if caller sent a filter). If empty set,
            # treat as "all" (coerce_types(None) gives all; empty only if bad input)
            if wanted_types:
                if code.startswith("PF") and "PF" not in wanted_types:
                    continue
                if code.startswith("ESI") and "ESI" not in wanted_types:
                    continue
                if code.startswith("PT") and "PT" not in wanted_types:
                    continue

            if code == "PF_EMP":
                pf_emp_amt = amt
            elif code == "PF_ER":
                # Split PF_ER into EPS/EPF 8.33/3.67 when positive; else whole to EPF
                if amt > 0:
                    eps = (amt * _as_decimal("0.833")).quantize(_as_decimal("0.01"))
                    epf = (amt - eps).quantize(_as_decimal("0.01"))
                    pf_eps_amt, pf_epf_amt = eps, epf
                else:
                    pf_epf_amt = amt
            elif code == "ESI_EMP":
                esi_emp_amt = amt
            elif code == "ESI_ER":
                esi_er_amt = amt
            elif code in ("PT_MH", "PT"):
                pt_amt = amt

        # Existing statutory lines
        existing = (
            db.session.query(PayRunItemLine, SalaryComponent)
            .join(SalaryComponent, PayRunItemLine.component_id == SalaryComponent.id)
            .filter(PayRunItemLine.item_id == it.id, PayRunItemLine.is_statutory.is_(True))
        ).all()
        existing_by_code = {sc.code: (line, sc) for (line, sc) in existing}

        def _upsert_line(component: SalaryComponent, amount):
            from decimal import Decimal
            amount = Decimal(str(amount or 0))
            existed = existing_by_code.get(component.code)
            if existed:
                line, _ = existed
                old = _as_decimal(line.amount or 0)
                line.amount = amount
                line.is_statutory = True
                return old, amount
            nl = PayRunItemLine(
                item_id=it.id, component_id=component.id, amount=amount, is_statutory=True
            )
            db.session.add(nl)
            return _as_decimal(0), amount

        changes = {}
        # Only upsert lines for requested types (if filter provided)
        if not wanted_types or "PF" in wanted_types:
            old, new = _upsert_line(comp_pf_emp, pf_emp_amt);  changes["PF_EMP"]    = {"old": _as_float(old), "new": _as_float(new)}
            old, new = _upsert_line(comp_pf_eps, pf_eps_amt);  changes["PF_ER_EPS"] = {"old": _as_float(old), "new": _as_float(new)}
            old, new = _upsert_line(comp_pf_epf, pf_epf_amt);  changes["PF_ER_EPF"] = {"old": _as_float(old), "new": _as_float(new)}
        if not wanted_types or "ESI" in wanted_types:
            old, new = _upsert_line(comp_esi_emp, esi_emp_amt);changes["ESI_EMP"]   = {"old": _as_float(old), "new": _as_float(new)}
            old, new = _upsert_line(comp_esi_er, esi_er_amt);  changes["ESI_ER"]    = {"old": _as_float(old), "new": _as_float(new)}
        if not wanted_types or "PT" in wanted_types:
            old, new = _upsert_line(comp_pt, pt_amt);          changes["PT"]        = {"old": _as_float(old), "new": _as_float(new)}

        # Update totals on item
        gross = _as_decimal(getattr(it, "gross", 0))
        d_pf = pf_emp_amt if (not wanted_types or "PF" in wanted_types) else _as_decimal(0)
        d_esi = esi_emp_amt if (not wanted_types or "ESI" in wanted_types) else _as_decimal(0)
        d_pt = pt_amt if (not wanted_types or "PT" in wanted_types) else _as_decimal(0)
        deductions = (d_pf + d_esi + d_pt).quantize(_as_decimal("0.01"))
        it.deductions = _as_float(deductions)
        it.net = _as_float((gross - deductions).quantize(_as_decimal("0.01")))

        # Audit trail (be robust: calc_meta may be list/dict/None)
        meta_raw = getattr(it, "calc_meta", None)
        meta = meta_raw if isinstance(meta_raw, dict) else {}
        audit = meta.get("audit") or []
        audit.append({
            "ts": datetime.utcnow().isoformat() + "Z",
            "changes": changes,
            "configs": cfg_ids,
        })
        meta["audit"] = audit
        # Preserve legacy non-dict payload if present
        if meta_raw is not None and not isinstance(meta_raw, dict):
            meta.setdefault("_legacy", meta_raw)
        it.calc_meta = meta

        applied += 1

    # Auto-bump calculated -> approved after applying lines (your existing behavior)
    if r.status == "calculated":
        r.status = "approved"
        if hasattr(r, "approved_at"):
            r.approved_at = datetime.utcnow()

    db.session.commit()
    return _ok({"items_updated": applied, "run": {"id": r.id, "status": r.status}})

# Determine which JSON column on PayRunItem stores components (fallback to any JSON field)
def _pick_item_components_field() -> Optional[str]:
    cols = list(PayRunItem.__table__.columns.items())
    # Prefer names containing component keywords and of JSON type
    for name, col in cols:
        nm = name.lower()
        if any(k in nm for k in ("components","component_json","breakup","breakdown")):
            if isinstance(col.type, _JSON):
                return name
    # Fallback: any JSON column
    for name, col in cols:
        if isinstance(col.type, _JSON):
            return name
    return None

ITEM_COMPONENTS_F = _pick_item_components_field()
