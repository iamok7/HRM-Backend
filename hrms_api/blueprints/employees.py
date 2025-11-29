from __future__ import annotations

from flask import Blueprint, request, jsonify
from sqlalchemy import or_, select
from datetime import datetime, date
from hrms_api.extensions import db
from hrms_api.models.employee import Employee
from hrms_api.models.master import Company, Location, Department, Designation, Grade, CostCenter
from flask_jwt_extended import jwt_required
from hrms_api.common.auth import requires_roles, requires_perms


# --- NEW IMPORTS (add near the top with your other imports) ---
import io, csv, re
from datetime import datetime, date, time, timedelta
from werkzeug.utils import secure_filename
from typing import Optional

# Attendance model — we’ll adapt to either .ts/.kind OR .punch_dt/.direction
try:
    from hrms_api.models.attendance import AttendancePunch
except Exception:
    AttendancePunch = None  # graceful fallback; we’ll fail with a clear error later
# --------------------------------------------------------------



# --- NEW: imports used by the tiny role attach helper ---
from hrms_api.models.user import User
from hrms_api.models.user import User
from hrms_api.models.security import Role, UserRole
from hrms_api.services.leave_policy_service import sync_balances_for_company_year
# --------------------------------------------------------

bp = Blueprint("employees", __name__, url_prefix="/api/v1/employees")

# ---------- envelopes ----------
def _ok(data=None, status=200, **meta):
    payload = {"success": True, "data": data}
    if meta: payload["meta"] = meta
    return jsonify(payload), status

def _fail(msg, status=400, code=None, detail=None):
    err = {"message": msg}
    if code: err["code"] = code
    if detail: err["detail"] = detail
    return jsonify({"success": False, "error": err}), status

# ---------- helpers ----------
def _int_arg(*names: str):
    """
    Return first present arg among names (camelCase/snake_case),
    cast to int. If present but invalid -> ValueError (422).
    """
    for n in names:
        if n in request.args:
            v = request.args.get(n)
            if v in (None, "", "null"): return None
            try:
                return int(v)
            except Exception:
                raise ValueError(f"{n} must be integer")
    return None

def _bool_arg(*names: str):
    for n in names:
        if n in request.args:
            v = (request.args.get(n) or "").lower()
            if v in ("true","1","yes"):  return True
            if v in ("false","0","no"):  return False
            raise ValueError(f"{n} must be true/false")
    return None

def _as_int(val, field_name):
    if val is None or val == "":
        return None
    try:
        return int(val)
    except Exception:
        raise ValueError(f"{field_name} must be integer")

def _page_limit():
    """
    Backward compatible:
      - page (default 1)
      - limit (default 20)  [original]
      - size (alias for limit)
    Clamped to [1,100]
    """
    # page
    try:
        page = max(int(request.args.get("page", 1)), 1)
    except Exception:
        page = 1
    # limit / size
    raw_limit = request.args.get("limit", None)
    raw_size  = request.args.get("size", None)
    use = raw_limit if raw_limit is not None else raw_size
    try:
        limit = int(use) if use is not None else 20
        limit = min(max(limit, 1), 100)
    except Exception:
        limit = 20
    return page, limit

def _parse_date(val):
    if not val: return None
    if isinstance(val, (date,)): return val
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try: return datetime.strptime(val, fmt).date()
        except Exception: pass
    return None

def _row(x: Employee):
    return {
        "id": x.id,
        "code": x.code,
        "email": x.email,
        "first_name": x.first_name,
        "last_name": x.last_name,
        "phone": x.phone,

        "company_id": x.company_id,
        "company_name": x.company.name if x.company else None,
        "location_id": x.location_id,
        "location_name": x.location.name if x.location else None,
        "department_id": x.department_id,
        "department_name": x.department.name if x.department else None,
        "designation_id": x.designation_id,
        "designation_name": x.designation.name if x.designation else None,
        "grade_id": x.grade_id,
        "grade_name": x.grade.name if x.grade else None,
        "cost_center_id": x.cost_center_id,
        "cost_center_code": x.cost_center.code if x.cost_center else None,

        "manager_id": x.manager_id,
        "manager_name": (
            (x.manager.first_name or "")
            + ((" " + x.manager.last_name) if x.manager and x.manager.last_name else "")
        ) if x.manager else None,

        "employment_type": x.employment_type,
        "status": x.status,
        "doj": x.doj.isoformat() if x.doj else None,
        "dol": x.dol.isoformat() if x.dol else None,
        "created_at": x.created_at.isoformat() if x.created_at else None,
    }

# ---------- NEW: role attach helpers (local to this file) ----------

def _ensure_user_has_role(user_id: int, role_code: str = "employee") -> bool:
    """
    Idempotently ensure the given user has `role_code`.
    Returns True if a new link is created, False if it already existed.
    """
    role = db.session.scalar(select(Role).where(Role.code == role_code))
    if not role:
        # Keep silent to avoid failing the whole request; caller may have not seeded RBAC yet.
        return False
    exists = db.session.scalar(
        select(UserRole).where(UserRole.user_id == user_id, UserRole.role_id == role.id)
    )
    if exists:
        return False
    db.session.add(UserRole(user_id=user_id, role_id=role.id))
    return True

def _attach_employee_role_if_user_exists(emp: Employee):
    """
    Try to find a corresponding User for this employee.
    - If found (by user_id or email), ensure they are linked (emp.user_id = user.id).
    - If NOT found, CREATE a new User with default password and link.
    - Attach 'employee' role to the user.
    """
    user = None
    uid = getattr(emp, "user_id", None)
    if uid:
        user = db.session.get(User, uid)
    
    if user is None and emp.email:
        user = db.session.scalar(select(User).where(User.email == emp.email.lower()))
        
    # Auto-create user if missing
    if user is None and emp.email:
        try:
            full_name = f"{emp.first_name} {emp.last_name or ''}".strip()
            user = User(email=emp.email.lower(), full_name=full_name, status='active')
            user.set_password("Welcome@123")
            db.session.add(user)
            db.session.flush() # get ID
        except Exception:
            # Fallback (e.g. race condition), just skip
            pass

    if user:
        # Link if not linked
        if not emp.user_id:
            emp.user_id = user.id
            db.session.add(emp) # Ensure update is tracked
            
        _ensure_user_has_role(user.id, "employee")

# ------------------------------------------------------------------

# ---------- routes ----------

# List (JWT only; no perm — matches your current working code)
@bp.get("")
@jwt_required()
def list_employees():
    q = Employee.query

    # filters (camelCase + snake_case aliases)
    try:
        cid = _int_arg("companyId", "company_id")
        did = _int_arg("deptId", "department_id")
        loc = _int_arg("locationId", "location_id")
        des = _int_arg("designationId", "designation_id")
        grd = _int_arg("gradeId", "grade_id")
    except ValueError as ex:
        return _fail(str(ex), 422)

    if cid: q = q.filter(Employee.company_id == cid)
    if did: q = q.filter(Employee.department_id == did)
    if loc: q = q.filter(Employee.location_id == loc)
    if des: q = q.filter(Employee.designation_id == des)
    if grd: q = q.filter(Employee.grade_id == grd)

    # status (string)
    status = (request.args.get("status") or "").strip()
    if status:
        q = q.filter(Employee.status == status.lower())

    # is_active (optional bool) — supports both `is_active` and `isActive`
    try:
        is_act = _bool_arg("is_active", "isActive")
        if is_act is True:
            q = q.filter(Employee.is_active.is_(True))
        elif is_act is False:
            q = q.filter(Employee.is_active.is_(False))
    except ValueError as ex:
        return _fail(str(ex), 422)

    # search
    s = (request.args.get("q") or "").strip()
    if s:
        like = f"%{s}%"
        q = q.filter(or_(Employee.code.ilike(like),
                         Employee.email.ilike(like),
                         Employee.first_name.ilike(like),
                         Employee.last_name.ilike(like)))

    page, limit = _page_limit()
    total = q.count()
    items = q.order_by(Employee.id.desc()).offset((page-1)*limit).limit(limit).all()
    return _ok([_row(i) for i in items], page=page, limit=limit, total=total)

# Get (JWT only)
@bp.get("/<int:eid>")
@jwt_required()
def get_employee(eid: int):
    x = Employee.query.get(eid)
    if not x: return _fail("Employee not found", 404)
    return _ok(_row(x))

# Create (perm + role: admin)  — keep your existing RBAC shape
# ------- Create (patched)
@bp.post("")
@requires_perms("employee.create")
@requires_roles("admin")
def create_employee():
    d = request.get_json(silent=True, force=True) or {}
    try:
        cid  = _as_int(d.get("company_id"),    "company_id")
        lid  = _as_int(d.get("location_id"),   "location_id")
        did  = _as_int(d.get("department_id"), "department_id")
        dsid = _as_int(d.get("designation_id"),"designation_id")
        gid  = _as_int(d.get("grade_id"),      "grade_id")
        ccid = _as_int(d.get("cost_center_id"),"cost_center_id")
        mid  = _as_int(d.get("manager_id"),    "manager_id")
    except ValueError as ex:
        return _fail(str(ex), 422)

    code  = (d.get("code") or "").strip()
    email = (d.get("email") or "").strip().lower()
    first = (d.get("first_name") or "").strip()

    # Auto-generate code if missing
    if not code:
        import time
        code = f"EMP-{int(time.time())}"

    if not (email and first and cid):
        return _fail("company_id, first_name, email are required", 422)

    # FK checks (now using proper ints)
    if not Company.query.get(cid): return _fail("Invalid company_id", 422)
    for fk, model, name in [
        (lid,  Location,    "location_id"),
        (did,  Department,  "department_id"),
        (dsid, Designation, "designation_id"),
        (gid,  Grade,       "grade_id"),
        (ccid, CostCenter,  "cost_center_id"),
        (mid,  Employee,    "manager_id"),
    ]:
        if fk and not model.query.get(fk): return _fail(f"Invalid {name}", 422)

    # Uniques
    if Employee.query.filter_by(company_id=cid, code=code).first():
        return _fail("Employee code already exists for this company", 409)
    if Employee.query.filter_by(email=email).first():
        return _fail("Email already exists", 409)

    x = Employee(
        company_id=cid,
        location_id=lid,
        department_id=did,
        designation_id=dsid,
        grade_id=gid,
        cost_center_id=ccid,
        manager_id=mid,
        code=code,
        email=email,
        first_name=first,
        last_name=(d.get("last_name") or "").strip() or None,
        phone=(d.get("phone") or "").strip() or None,
        employment_type=(d.get("employment_type") or "fulltime").lower(),
        status=(d.get("status") or "active").lower(),
        doj=_parse_date(d.get("doj")),
        dol=_parse_date(d.get("dol")),
    )

    # --- important: single transaction; attach role if a matching user exists
    db.session.add(x)
    db.session.flush()  # x.id available; if a user_id is set by triggers/logic, we can see it

    _attach_employee_role_if_user_exists(x)  # <--- ensures 'employee' role is linked to the User (and creates user if missing)

    db.session.commit()
    
    # Sync leave balances
    try:
        sync_balances_for_company_year(
            company_id=x.company_id,
            year=datetime.utcnow().year,
            employee_ids=[x.id]
        )
    except Exception as e:
        # Log error but don't fail the request
        pass
        
    return _ok(_row(x), status=201)


# Update (perm + role: admin)
# ------- Update (patched)
@bp.put("/<int:eid>")
@requires_perms("employee.update")
@requires_roles("admin")
def update_employee(eid: int):
    x = Employee.query.get(eid)
    if not x: return _fail("Employee not found", 404)
    d = request.get_json(silent=True, force=True) or {}

    # company_id (coerce to int if present)
    if "company_id" in d:
        try:
            cid = _as_int(d.get("company_id"), "company_id")
        except ValueError as ex:
            return _fail(str(ex), 422)
        if not Company.query.get(cid): return _fail("Invalid company_id", 422)
        x.company_id = cid

    # code (uniqueness within company)
    if "code" in d:
        new_code = (d["code"] or "").strip()
        if not new_code: return _fail("code cannot be empty", 422)
        company_for_check = x.company_id
        if "company_id" in d:
            try:
                company_for_check = _as_int(d.get("company_id"), "company_id")
            except ValueError as ex:
                return _fail(str(ex), 422)
        if Employee.query.filter(
            Employee.id != eid,
            Employee.company_id == company_for_check,
            Employee.code == new_code
        ).first():
            return _fail("Employee code already exists for this company", 409)
        x.code = new_code

    # email
    if "email" in d:
        new_email = (d["email"] or "").strip().lower()
        if not new_email: return _fail("email cannot be empty", 422)
        if Employee.query.filter(Employee.id != eid, Employee.email == new_email).first():
            return _fail("Email already exists", 409)
        x.email = new_email

    # FK fields (coerce to ints)
    for key, model in [
        ("location_id",    Location),
        ("department_id",  Department),
        ("designation_id", Designation),
        ("grade_id",       Grade),
        ("cost_center_id", CostCenter),
        ("manager_id",     Employee),
    ]:
        if key in d:
            try:
                val = _as_int(d.get(key), key) if d.get(key) is not None else None
            except ValueError as ex:
                return _fail(str(ex), 422)
            if val and not model.query.get(val): return _fail(f"Invalid {key}", 422)
            setattr(x, key, val)

    for key in ("first_name", "last_name", "phone", "employment_type", "status"):
        if key in d: setattr(x, key, (d[key] or "").strip() or None)
    if "doj" in d: x.doj = _parse_date(d["doj"])
    if "dol" in d: x.dol = _parse_date(d["dol"])

    db.session.commit()
    return _ok(_row(x))


# Delete (soft) (perm + role: admin)
@bp.delete("/<int:eid>")
@requires_perms("employee.delete")
@requires_roles("admin")
def delete_employee(eid: int):
    x = Employee.query.get(eid)
    if not x: return _fail("Employee not found", 404)
    x.status = "inactive"
    if not x.dol: x.dol = date.today()
    db.session.commit()
    return _ok({"id": eid, "status": "inactive", "dol": x.dol.isoformat()})


# ===== Attendance import (robust, with ORM + reflection fallback) =====
import io, csv, re
from datetime import datetime, date, time
from werkzeug.utils import secure_filename
from sqlalchemy import select, and_, MetaData, Table
from sqlalchemy.orm import class_mapper, ColumnProperty

# Try common model names (ORM path)
AttendancePunch = None
AttendanceDay = None
try:
    from hrms_api.models.attendance import AttendancePunch as _AP
    AttendancePunch = _AP
except Exception:
    pass

if AttendancePunch is None:
    try:
        from hrms_api.models.attendance import Attendance as _AD
        AttendanceDay = _AD
    except Exception:
        try:
            from hrms_api.models.attendance import DailyAttendance as _DA
            AttendanceDay = _DA
        except Exception:
            AttendanceDay = None

_ALLOWED_EXTS = {".csv", ".xlsx"}
from datetime import datetime

def _default_cols_for_punch(tbl) -> dict:
    """
    Build default values for commonly NOT NULL punch columns if present.
    We only fill what exists. Safe for both reflection Table and ORM model.__table__.
    """
    cols = {c.name: c for c in tbl.columns}
    now = datetime.utcnow()
    vals = {}
    if "source" in cols:
        vals["source"] = "import"
    if "created_at" in cols:
        vals["created_at"] = now
    if "updated_at" in cols:
        vals["updated_at"] = now
    if "method" in cols:
        vals["method"] = "import"
    if "device_id" in cols:
        # if device_id is NOT NULL, give a neutral marker
        if cols["device_id"].nullable is False:
            vals["device_id"] = "IMPORT"
    if "note" in cols:
        # optional note to help audit
        vals.setdefault("note", "bulk-import")
    return vals


def _emp_info(emp: Employee) -> dict:
    name = " ".join([p for p in [(emp.first_name or "").strip(), (emp.last_name or "").strip()] if p]) or None
    return {
        "id": emp.id,
        "code": emp.code,
        "name": name,
        "email": emp.email
    }

# ---------- file utils ----------
def _ext_from_filename(fn: str) -> str:
    fn = (fn or "").lower().strip()
    m = re.search(r"\.[a-z0-9]+$", fn)
    return m.group(0) if m else ""

def _iter_rows_from_csv(file_storage) -> list[dict]:
    # read bytes → decode → StringIO (avoids SpooledTemporaryFile readable() issue)
    try:
        file_storage.stream.seek(0)
    except Exception:
        pass
    raw = file_storage.read()
    if isinstance(raw, bytes):
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
    else:
        text = str(raw)
    sio = io.StringIO(text)
    reader = csv.DictReader(sio)
    rows = []
    for row in reader:
        norm = {}
        for k, v in row.items():
            key = (k or "").strip().lower()
            if isinstance(v, str):
                v = v.strip()
            norm[key] = v
        rows.append(norm)
    return rows

def _iter_rows_from_xlsx(file_storage) -> list[dict]:
    try:
        import openpyxl
    except Exception:
        raise RuntimeError("XLSX support requires 'openpyxl'. Install it or upload CSV.")
    try:
        file_storage.stream.seek(0)
    except Exception:
        pass
    data = file_storage.read()
    if not isinstance(data, (bytes, bytearray)):
        data = str(data).encode("utf-8")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            headers = [str(h).strip().lower() if h is not None else "" for h in row]
            continue
        rec = {}
        for j, val in enumerate(row):
            key = headers[j] if j < len(headers) else f"col{j+1}"
            rec[key] = val if val is not None else ""
        rows.append(rec)
    return rows

# ---------- parsing ----------
def _parse_time_like(val) -> time | None:
    if val is None:
        return None
    if isinstance(val, time):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except Exception:
            pass
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S",
                "%d-%m-%Y %H:%M", "%d-%m-%Y %H:%M:%S",
                "%Y/%m/%d %H:%M", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).time()
        except Exception:
            pass
    return None

def _parse_date_like(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

# ---------- employee lookup ----------
def _find_employee_for_row(row: dict) -> Employee | None:
    emp = None
    eid = row.get("employee_id") or row.get("id")
    if eid:
        try:
            emp = Employee.query.get(int(eid))
            if emp and (emp.status or "").lower() != "inactive":
                return emp
        except Exception:
            pass
    code = (row.get("employee_code") or row.get("code") or "").strip()
    if code:
        emp = Employee.query.filter(
            Employee.code == code,
            Employee.status != "inactive"
        ).first()
        if emp:
            return emp
    email = (row.get("email") or "").strip().lower()
    if email:
        emp = Employee.query.filter(
            Employee.email == email,
            Employee.status != "inactive"
        ).first()
        if emp:
            return emp
    return None

# ---------- ORM model detection ----------
def _columns_of(model) -> set[str]:
    try:
        return {prop.key for prop in class_mapper(model).iterate_properties if isinstance(prop, ColumnProperty)}
    except Exception:
        return set()

def _detect_punch_model():
    candidates_ts = ["ts", "punch_dt", "punch_time", "timestamp", "time", "clock_at", "datetime"]
    candidates_kind = ["kind", "direction", "type", "punch_type", "action", "io"]
    model = AttendancePunch
    if model is None:
        return None
    cols = _columns_of(model)
    ts_field = next((c for c in candidates_ts if c in cols), None)
    kind_field = next((c for c in candidates_kind if c in cols), None)
    if ts_field and kind_field and "employee_id" in cols:
        return model, ts_field, kind_field
    return None

def _detect_day_model():
    model = AttendanceDay
    if model is None:
        return None
    cols = _columns_of(model)
    date_field = next((c for c in ["work_date", "date", "att_date"] if c in cols), None)
    in_field   = next((c for c in ["in_time", "in", "time_in", "punch_in"] if c in cols), None)
    out_field  = next((c for c in ["out_time", "out", "time_out", "punch_out"] if c in cols), None)
    if date_field and in_field and out_field and "employee_id" in cols:
        return model, date_field, in_field, out_field
    return None

# ---------- DB reflection fallback ----------
_PUNCH_TS_CANDS = ["ts", "punch_dt", "punch_time", "timestamp", "datetime", "time", "clock_at"]
_PUNCH_KIND_CANDS = ["kind", "direction", "type", "punch_type", "action", "io"]
_DAY_DATE_CANDS = ["work_date", "date", "att_date"]
_DAY_IN_CANDS = ["in_time", "time_in", "in", "punch_in"]
_DAY_OUT_CANDS = ["out_time", "time_out", "out", "punch_out"]
_TABLE_NAME_HINTS = [
    "attendance_punch", "attendance_punches", "punch", "punches",
    "attendance_day", "attendance_days", "daily_attendance", "attendance"
]

def _reflect_tables():
    md = MetaData()
    md.reflect(bind=db.engine)
    return md.tables  # dict name->Table

def _reflect_punch_table():
    tables = _reflect_tables()
    # Prefer hinted names first
    names = list(tables.keys())
    hinted = [n for n in names if any(h in n.lower() for h in _TABLE_NAME_HINTS)]
    scan_order = hinted + [n for n in names if n not in hinted]
    for name in scan_order:
        tbl: Table = tables[name]
        cols = {c.name.lower(): c for c in tbl.columns}
        if "employee_id" not in cols:
            continue
        ts_col = next((cols[c] for c in _PUNCH_TS_CANDS if c in cols), None)
        kind_col = next((cols[c] for c in _PUNCH_KIND_CANDS if c in cols), None)
        if ts_col is not None and kind_col is not None:
            return tbl, ts_col.name, kind_col.name
    return None

def _reflect_day_table():
    tables = _reflect_tables()
    names = list(tables.keys())
    hinted = [n for n in names if any(h in n.lower() for h in _TABLE_NAME_HINTS)]
    scan_order = hinted + [n for n in names if n not in hinted]
    for name in scan_order:
        tbl: Table = tables[name]
        cols = {c.name.lower(): c for c in tbl.columns}
        if "employee_id" not in cols:
            continue
        date_col = next((cols[c] for c in _DAY_DATE_CANDS if c in cols), None)
        in_col   = next((cols[c] for c in _DAY_IN_CANDS   if c in cols), None)
        out_col  = next((cols[c] for c in _DAY_OUT_CANDS  if c in cols), None)
        if date_col and in_col and out_col:
            return tbl, date_col.name, in_col.name, out_col.name
    return None

def _ensure_unique_punch_orm(model, emp_id: int, stamp: datetime, kind: str, ts_field: str, kind_field: str) -> bool:
    exists = db.session.scalar(
        select(model).where(
            getattr(model, "employee_id") == emp_id,
            getattr(model, ts_field) == stamp,
            getattr(model, kind_field) == kind
        )
    )
    if exists:
        return False
    payload = { "employee_id": emp_id, ts_field: stamp, kind_field: kind }
    now = datetime.utcnow()
    # set common defaults if those columns exist on the ORM model
    for col, val in {
        "source": "import",
        "created_at": now,
        "updated_at": now,
        "method": "import",
        "note": "bulk-import"
    }.items():
        if hasattr(model, col):
            payload[col] = val
    if hasattr(model, "device_id"):
        payload.setdefault("device_id", "IMPORT")
    obj = model(**payload)
    db.session.add(obj)
    return True


def _ensure_unique_punch_core(tbl, emp_id: int, stamp: datetime, kind: str, ts_col: str, kind_col: str) -> bool:
    exists = db.session.execute(
        select(tbl).where(
            tbl.c.employee_id == emp_id,
            getattr(tbl.c, ts_col) == stamp,
            getattr(tbl.c, kind_col) == kind
        )
    ).first()
    if exists:
        return False
    vals = {"employee_id": emp_id, ts_col: stamp, kind_col: kind}
    vals.update(_default_cols_for_punch(tbl))
    db.session.execute(tbl.insert().values(**vals))
    return True


# ---------- route ----------
@bp.post("/attendance/import")
@requires_perms("attendance.punch.create")   # Option 1 you chose
@requires_roles("admin")
def import_attendance_file():
    """
    Multipart form-data with 'file' (CSV/XLSX).
    Columns:
      - one of: employee_id | employee_code | email
      - date: YYYY-MM-DD (also accepts DD-MM-YYYY / YYYY/MM/DD)
      - in_time, out_time: HH:MM (or HH:MM:SS / 9:30 AM)
    Works with ORM models OR reflected tables.
    """
    if "file" not in request.files:
        return _fail("Upload a file in form field 'file'", 422)

    f = request.files["file"]
    fname = secure_filename(f.filename or "")
    ext = _ext_from_filename(fname)
    if ext not in _ALLOWED_EXTS:
        return _fail("Unsupported file type. Upload .csv or .xlsx", 415)

    # Parse rows
    try:
        rows = _iter_rows_from_csv(f) if ext == ".csv" else _iter_rows_from_xlsx(f)
    except RuntimeError as e:
        return _fail(str(e), 415)
    except Exception as e:
        return _fail(f"Failed to read file: {e}", 400)

    # Detect schema: ORM first, then reflection
    punch_model = _detect_punch_model()
    day_model   = _detect_day_model()
    punch_tbl   = None
    day_tbl     = None

    if not punch_model and not day_model:
        punch_tbl = _reflect_punch_table()
        if not punch_tbl:
            day_tbl = _reflect_day_table()

    if not (punch_model or day_model or punch_tbl or day_tbl):
        return _fail("Attendance schema not recognized. Expect either a punch table (ts+kind) or a day table (date+in_time+out_time).", 500)

    accepted = 0
    created  = 0
    updated  = 0
    errors   = []
    present_rows: list[dict] = []
    absent_rows: list[dict] = []
    absent = 0
    for idx, row in enumerate(rows, start=2):
        try:
            emp = _find_employee_for_row(row)
            if not emp:
                errors.append({"row": idx, "error": "employee not found (id/code/email)", "row_data": row})
                continue

            d = _parse_date_like(row.get("date"))
            if not d:
                errors.append({"row": idx, "error": "invalid/missing date", "row_data": row})
                continue

            t_in  = _parse_time_like(row.get("in_time"))
            t_out = _parse_time_like(row.get("out_time"))
            # skip silently but track as absent for summary
            if not t_in and not t_out:
                # (optional) collect absent count in memory for meta
                absent += 1
                absent_rows.append({
                    "employee": _emp_info(emp),
                    "date": d.isoformat()
                })
                continue


            accepted += 1

            if punch_model:
                model, ts_field, kind_field = punch_model
                if t_in:
                    dt_in = datetime.combine(d, t_in)
                    if _ensure_unique_punch_orm(model, emp.id, dt_in, "in", ts_field, kind_field):
                        created += 1
                if t_out:
                    dt_out = datetime.combine(d, t_out)
                    if _ensure_unique_punch_orm(model, emp.id, dt_out, "out", ts_field, kind_field):
                        created += 1

            elif day_model:
                model, date_field, in_field, out_field = day_model
                existing = db.session.scalar(
                    select(model).where(
                        getattr(model, "employee_id") == emp.id,
                        getattr(model, date_field) == d
                    )
                )
                if existing:
                    if t_in:  setattr(existing, in_field,  t_in)
                    if t_out: setattr(existing, out_field, t_out)
                    updated += 1
                else:
                    obj = model(employee_id=emp.id, **{
                        date_field: d,
                        in_field:   t_in,
                        out_field:  t_out
                    })
                    db.session.add(obj)
                    created += 1

            elif punch_tbl:
                tbl, ts_col, kind_col = punch_tbl
                if t_in:
                    dt_in = datetime.combine(d, t_in)
                    if _ensure_unique_punch_core(tbl, emp.id, dt_in, "in", ts_col, kind_col):
                        created += 1
                if t_out:
                    dt_out = datetime.combine(d, t_out)
                    if _ensure_unique_punch_core(tbl, emp.id, dt_out, "out", ts_col, kind_col):
                        created += 1

            elif day_tbl:
                tbl, date_col, in_col, out_col = day_tbl
                # upsert (employee_id, date)
                existing = db.session.execute(
                    select(tbl).where(
                        tbl.c.employee_id == emp.id,
                        getattr(tbl.c, date_col) == d
                    )
                ).first()

                # record present row for UI table
                present_rows.append({
                    "employee": _emp_info(emp),
                    "date": d.isoformat(),
                    "in_time": t_in.isoformat() if t_in else None,
                    "out_time": t_out.isoformat() if t_out else None
                })

                if existing:
                    upd_vals = {}
                    if t_in:  upd_vals[in_col]  = t_in
                    if t_out: upd_vals[out_col] = t_out
                    if upd_vals:
                        db.session.execute(tbl.update().where(
                            tbl.c.employee_id == emp.id,
                            getattr(tbl.c, date_col) == d
                        ).values(**upd_vals))
                        updated += 1
                else:
                    db.session.execute(tbl.insert().values(
                        employee_id=emp.id,
                        **{date_col: d, in_col: t_in, out_col: t_out}
                    ))
                    created += 1

        except Exception as ex:
            db.session.rollback()
            errors.append({"row": idx, "error": f"unexpected: {ex}", "row_data": row})

    try:
        db.session.commit()
    except Exception as ex:
        db.session.rollback()
        return _fail(f"DB commit failed: {ex}", 500)
    
    from collections import defaultdict
    by_date = defaultdict(lambda: {"present": 0, "absent": 0})
    for r in present_rows:
        by_date[r["date"]]["present"] += 1
    for r in absent_rows:
        by_date[r["date"]]["absent"] += 1
    summary_by_date = dict(by_date)

    return _ok(
        data={
            "accepted_rows": accepted,
            "created": created,
            "updated": updated,
            "present_count": len(present_rows),
            "absent_count": absent,
            "present_rows": present_rows,   # full detail for UI table
            "absent_rows": absent_rows,     # who had no punches
            "summary_by_date": summary_by_date
        },
        status=200,
        total_rows=len(rows)
    )


@bp.get("/attendance/daily")
@requires_perms("attendance.read")
@requires_roles("admin", "hr", "manager")
@jwt_required()
def attendance_daily():
    d = _parse_date_like(request.args.get("date"))
    if not d:
        return _fail("Provide date=YYYY-MM-DD", 422)

    # filters
    q = (request.args.get("q") or "").strip()
    try:
        cid = _int_arg("company_id", "companyId")
        did = _int_arg("department_id", "deptId")
        lid = _int_arg("location_id", "locationId")
    except ValueError as ex:
        return _fail(str(ex), 422)

    include_absent = True
    try:
        ia = _bool_arg("include_absent")
        if ia is not None:
            include_absent = ia
    except ValueError as ex:
        return _fail(str(ex), 422)

    # base employees set (active only)
    equery = Employee.query.filter(Employee.status != "inactive")
    if cid: equery = equery.filter(Employee.company_id == cid)
    if did: equery = equery.filter(Employee.department_id == did)
    if lid: equery = equery.filter(Employee.location_id == lid)
    if q:
        like = f"%{q}%"
        equery = equery.filter(
            or_(Employee.code.ilike(like),
                Employee.email.ilike(like),
                Employee.first_name.ilike(like),
                Employee.last_name.ilike(like))
        )

    # paging over employees (we’ll compute their attendance per employee)
    page, limit = _page_limit()
    total = equery.count()
    employees = (equery
                 .order_by(Employee.id.asc())
                 .offset((page-1)*limit)
                 .limit(limit)
                 .all())

    # detect attendance schema (reusing your importer helpers)
    punch_model = _detect_punch_model()
    day_model   = _detect_day_model()
    punch_tbl   = None
    day_tbl     = None
    if not (punch_model or day_model):
        punch_tbl = _reflect_punch_table()
        if not punch_tbl:
            day_tbl = _reflect_day_table()

    # compute window
    start_dt = datetime(d.year, d.month, d.day, 0, 0, 0)
    end_dt   = datetime(d.year, d.month, d.day, 23, 59, 59)

    rows = []
    # ---- using punch model/table: present if any punch between start_dt..end_dt
    if punch_model or punch_tbl:
        # ORM path
        if punch_model:
            model, ts_field, kind_field = punch_model
            for emp in employees:
                # pull all punches for the day (light weight per page)
                punches = (db.session.query(model)
                           .filter(getattr(model, "employee_id") == emp.id,
                                   getattr(model, ts_field) >= start_dt,
                                   getattr(model, ts_field) <= end_dt)
                           .all())
                if punches:
                    # first IN / last OUT if we know kind; else min/max ts
                    in_ts = None
                    out_ts = None
                    if kind_field:
                        ins  = [getattr(p, ts_field) for p in punches if (getattr(p, kind_field) or "").lower() == "in"]
                        outs = [getattr(p, ts_field) for p in punches if (getattr(p, kind_field) or "").lower() == "out"]
                        in_ts  = min(ins)  if ins  else min(getattr(p, ts_field) for p in punches)
                        out_ts = max(outs) if outs else max(getattr(p, ts_field) for p in punches)
                    else:
                        in_ts  = min(getattr(p, ts_field) for p in punches)
                        out_ts = max(getattr(p, ts_field) for p in punches)
                    rows.append({
                        "employee": _emp_info(emp),
                        "date": d.isoformat(),
                        "status": "present",
                        "in_time": in_ts.time().isoformat(timespec="seconds"),
                        "out_time": out_ts.time().isoformat(timespec="seconds")
                    })
                elif include_absent:
                    rows.append({"employee": _emp_info(emp), "date": d.isoformat(), "status": "absent",
                                 "in_time": None, "out_time": None})
        else:
            # Core/reflection path
            tbl, ts_col, kind_col = punch_tbl
            for emp in employees:
                punches = db.session.execute(
                    select(tbl).where(
                        tbl.c.employee_id == emp.id,
                        getattr(tbl.c, ts_col) >= start_dt,
                        getattr(tbl.c, ts_col) <= end_dt
                    )
                ).all()
                if punches:
                    # build min/max from rows
                    ts_values = [getattr(r[0], ts_col) if hasattr(r[0], ts_col) else r[0][ts_col] for r in punches] \
                                if hasattr(punches[0][0], ts_col) else [r._mapping[ts_col] for r in punches]
                    in_ts = min(ts_values)
                    out_ts = max(ts_values)
                    rows.append({"employee": _emp_info(emp), "date": d.isoformat(), "status": "present",
                                 "in_time": in_ts.time().isoformat(timespec="seconds"),
                                 "out_time": out_ts.time().isoformat(timespec="seconds")})
                elif include_absent:
                    rows.append({"employee": _emp_info(emp), "date": d.isoformat(), "status": "absent",
                                 "in_time": None, "out_time": None})

    # ---- using day model/table: present if (in_time or out_time) or status != 'absent'
    elif day_model or day_tbl:
        if day_model:
            model, date_field, in_field, out_field = day_model
            for emp in employees:
                rec = db.session.scalar(
                    select(model).where(
                        getattr(model, "employee_id") == emp.id,
                        getattr(model, date_field) == d
                    )
                )
                if rec:
                    in_t  = getattr(rec, in_field, None)
                    out_t = getattr(rec, out_field, None)
                    status = getattr(rec, "status", None)
                    present = (in_t is not None) or (out_t is not None) or (status and status.lower() != "absent")
                    rows.append({
                        "employee": _emp_info(emp),
                        "date": d.isoformat(),
                        "status": "present" if present else "absent",
                        "in_time": in_t.isoformat() if hasattr(in_t, "isoformat") and in_t else None,
                        "out_time": out_t.isoformat() if hasattr(out_t, "isoformat") and out_t else None
                    })
                elif include_absent:
                    rows.append({"employee": _emp_info(emp), "date": d.isoformat(), "status": "absent",
                                 "in_time": None, "out_time": None})
        else:
            tbl, date_col, in_col, out_col = day_tbl
            for emp in employees:
                rec = db.session.execute(
                    select(tbl).where(
                        tbl.c.employee_id == emp.id,
                        getattr(tbl.c, date_col) == d
                    )
                ).first()
                if rec:
                    row = rec._mapping
                    in_t  = row.get(in_col)
                    out_t = row.get(out_col)
                    status = row.get("status")
                    present = (in_t is not None) or (out_t is not None) or (status and str(status).lower() != "absent")
                    rows.append({
                        "employee": _emp_info(emp),
                        "date": d.isoformat(),
                        "status": "present" if present else "absent",
                        "in_time": in_t.isoformat() if hasattr(in_t, "isoformat") and in_t else None,
                        "out_time": out_t.isoformat() if hasattr(out_t, "isoformat") and out_t else None
                    })
                elif include_absent:
                    rows.append({"employee": _emp_info(emp), "date": d.isoformat(), "status": "absent",
                                 "in_time": None, "out_time": None})

    else:
        return _fail("Attendance schema not recognized (no punch/day table).", 500)

    # quick per-day counts (for the page we returned)
    present_cnt = sum(1 for r in rows if r["status"] == "present")
    absent_cnt  = sum(1 for r in rows if r["status"] == "absent")

    return _ok(rows, page=page, limit=limit, total=total,
               date=d.isoformat(), present=present_cnt, absent=absent_cnt)

# ---------- SELF ATTENDANCE (employee view) ----------
@bp.get("/attendance/me")
@requires_roles("employee")                 # only employees
@requires_perms("attendance.read")
@jwt_required()
def attendance_me():
    """
    Show MY attendance for a given date (or today if omitted).
    Resolves current user -> Employee via Employee.user_id.
    Query: ?date=YYYY-MM-DD
    """
    from flask_jwt_extended import get_jwt_identity
    uid = get_jwt_identity()

    # find my employee record
    me = Employee.query.filter(Employee.user_id == uid).first()
    if not me:
        return _fail("No employee linked to this user", 404)

    # which day?
    d = _parse_date_like(request.args.get("date")) or date.today()

    # detect attendance schema (same as daily endpoint)
    punch_model = _detect_punch_model()
    day_model   = _detect_day_model()
    punch_tbl   = None
    day_tbl     = None
    if not (punch_model or day_model):
        punch_tbl = _reflect_punch_table()
        if not punch_tbl:
            day_tbl = _reflect_day_table()

    start_dt = datetime(d.year, d.month, d.day, 0, 0, 0)
    end_dt   = datetime(d.year, d.month, d.day, 23, 59, 59)

    # build one-row snapshot
    if punch_model:
        model, ts_field, kind_field = punch_model
        punches = (db.session.query(model)
                   .filter(getattr(model, "employee_id") == me.id,
                           getattr(model, ts_field) >= start_dt,
                           getattr(model, ts_field) <= end_dt)
                   .all())
        if punches:
            ins  = [getattr(p, ts_field) for p in punches if (getattr(p, kind_field) or "").lower() == "in"] if kind_field else []
            outs = [getattr(p, ts_field) for p in punches if (getattr(p, kind_field) or "").lower() == "out"] if kind_field else []
            in_ts  = (min(ins) if ins else min(getattr(p, ts_field) for p in punches)).time().isoformat(timespec="seconds")
            out_ts = (max(outs) if outs else max(getattr(p, ts_field) for p in punches)).time().isoformat(timespec="seconds")
            return _ok({
                "employee": _emp_info(me),
                "date": d.isoformat(),
                "status": "present",
                "in_time": in_ts,
                "out_time": out_ts
            })
        return _ok({"employee": _emp_info(me), "date": d.isoformat(), "status": "absent", "in_time": None, "out_time": None})

    if punch_tbl:
        tbl, ts_col, kind_col = punch_tbl
        punches = db.session.execute(
            select(tbl).where(
                tbl.c.employee_id == me.id,
                getattr(tbl.c, ts_col) >= start_dt,
                getattr(tbl.c, ts_col) <= end_dt
            )
        ).all()
        if punches:
            # reflect rows → min/max ts
            ts_values = [r._mapping[ts_col] for r in punches]
            in_ts  = min(ts_values).time().isoformat(timespec="seconds")
            out_ts = max(ts_values).time().isoformat(timespec="seconds")
            return _ok({"employee": _emp_info(me), "date": d.isoformat(), "status": "present",
                        "in_time": in_ts, "out_time": out_ts})
        return _ok({"employee": _emp_info(me), "date": d.isoformat(), "status": "absent", "in_time": None, "out_time": None})

    if day_model:
        model, date_field, in_field, out_field = day_model
        rec = db.session.scalar(select(model).where(
            getattr(model, "employee_id") == me.id,
            getattr(model, date_field) == d
        ))
        if rec:
            in_t  = getattr(rec, in_field, None)
            out_t = getattr(rec, out_field, None)
            status = getattr(rec, "status", None)
            present = (in_t is not None) or (out_t is not None) or (status and str(status).lower() != "absent")
            return _ok({
                "employee": _emp_info(me), "date": d.isoformat(),
                "status": "present" if present else "absent",
                "in_time": in_t.isoformat() if hasattr(in_t, "isoformat") and in_t else None,
                "out_time": out_t.isoformat() if hasattr(out_t, "isoformat") and out_t else None
            })
        return _ok({"employee": _emp_info(me), "date": d.isoformat(), "status": "absent", "in_time": None, "out_time": None})

    if day_tbl:
        tbl, date_col, in_col, out_col = day_tbl
        rec = db.session.execute(select(tbl).where(
            tbl.c.employee_id == me.id,
            getattr(tbl.c, date_col) == d
        )).first()
        if rec:
            row = rec._mapping
            in_t  = row.get(in_col)
            out_t = row.get(out_col)
            status = row.get("status")
            present = (in_t is not None) or (out_t is not None) or (status and str(status).lower() != "absent")
            return _ok({
                "employee": _emp_info(me), "date": d.isoformat(),
                "status": "present" if present else "absent",
                "in_time": in_t.isoformat() if hasattr(in_t, "isoformat") and in_t else None,
                "out_time": out_t.isoformat() if hasattr(out_t, "isoformat") and out_t else None
            })
        return _ok({"employee": _emp_info(me), "date": d.isoformat(), "status": "absent", "in_time": None, "out_time": None})
