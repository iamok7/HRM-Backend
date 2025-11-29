import requests
import openpyxl
import io
import os

BASE_URL = "http://localhost:5001/api/v1"
EMAIL = "admin@hrms.local"
PASSWORD = "password"

def verify_excel():
    session = requests.Session()
    
    # 1. Login
    print("Logging in...")
    resp = session.post(f"{BASE_URL}/auth/login", json={"email": EMAIL, "password": PASSWORD})
    if resp.status_code != 200:
        print("Login failed:", resp.text)
        return
    token = resp.json()["access"]
    headers = {"Authorization": f"Bearer {token}"}
    
    # 2. Run Report (ID 2)
    print("Running PAYROLL_REGISTER report...")
    resp = session.post(
        f"{BASE_URL}/rgs/reports/2/run",
        headers=headers,
        json={"params": {"company_id": 13, "year": 2025, "month": 11}}
    )
    if resp.status_code != 200:
        print("Run failed:", resp.text)
        return
    
    run_data = resp.json()["data"]
    output_id = run_data["outputs"][0]["id"]
    print(f"Report run successful. Output ID: {output_id}")
    
    # 3. Download Output
    print("Downloading output...")
    resp = session.get(f"{BASE_URL}/rgs/outputs/{output_id}/download", headers=headers)
    
    if resp.status_code != 200:
        print("Download failed:", resp.text)
        return
        
    # Check Headers
    print("Content-Type:", resp.headers.get("Content-Type"))
    print("Content-Disposition:", resp.headers.get("Content-Disposition"))
    
    # 4. Verify Excel Content
    try:
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        
        # Get headers (first row)
        excel_headers = [cell.value for cell in ws[1]]
        print("\nExcel Headers Found:")
        print(excel_headers)
        
        # Check for key columns
        required = ["emp_code", "employee_name", "basic", "hra", "pf_employee", "net_pay", "bank_account_number"]
        missing = [col for col in required if col not in excel_headers]
        
        if missing:
            print(f"\nFAILED: Missing columns: {missing}")
        else:
            print("\nSUCCESS: All key columns present.")
            
        print(f"Total Rows: {ws.max_row}")
        
    except Exception as e:
        print(f"Failed to parse Excel: {e}")

if __name__ == "__main__":
    verify_excel()
